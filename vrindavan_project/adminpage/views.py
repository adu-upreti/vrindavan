from django.shortcuts import render,HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from .forms import *
from .models import *
from logsign.models import *
from .forms import RestaurantInfoForm
import openpyxl

@login_required
def admin_dashboard(request):
    team_list = Add_Team.objects.all()
    
    data = {
        'dashboard_active_page': 'active',
        'teamlist': team_list,  
    }
    return render(request, "adminfile/dashboard.html", data)

@login_required
def Product_form(request):
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Product is saved successfully!")
            return redirect('product_form')
        else:
            messages.error(request, "There was an error saving the product.")
    else:
        form = ProductForm()

    return render(request, "adminfile/add-product.html", {"form": form})


@login_required
def update_product(request, product_id):
    product = get_object_or_404(Products, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('admin_product')
    else:
        form = ProductForm(instance=product)
    return render(request, 'adminfile/update_product.html', {'form': form})


@login_required
def delete_product(request, product_id):
    product = get_object_or_404(Products, id=product_id)
    product.delete()
    messages.success(request, 'Product has been deleted successfully.')
    return redirect('admin_product')


@login_required
def delete_selected_products(request):
    if request.method == 'POST':
        selected_products = request.POST.getlist('selected_products')
        if selected_products:
            Products.objects.filter(id__in=selected_products).delete()
            messages.success(request, 'Selected products have been deleted successfully.')
        else:
            messages.error(request, 'No products were selected for deletion.')
    return redirect('admin_product')
   

@login_required
def add_cat(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_product')
    else:
        form = CategoryForm()
    return render(request, "adminfile/add-category.html", {'form': form})

@login_required
def delete_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    category.delete()
    return redirect('admin_product')


@login_required
def A_product(request):
    category_list = Category.objects.all()
    query = request.GET.get('q')

    if query:
        product_list =  Products.objects.filter(
            Q(name__icontains=query)
        ).order_by('name')

    else:

        product_list = Products.objects.all()
    

    datas = {
        'categories': category_list,
        'products': product_list,
        'query': query,
    }

    return render(request, "adminfile/products.html", datas)


@login_required
def Userlist(request):
    query = request.GET.get('q')

    if query:
        user_list = Profile.objects.filter(
            Q(full_name__icontains=query) | 
            Q(user__email__icontains=query)
        ).order_by('full_name')
    else:
        user_list = Profile.objects.all().order_by('full_name')

    # Check if the export button was clicked
    if 'export' in request.GET:
        return export_users_to_excel(user_list)

    datas = {
        'userlist': user_list,
        'query': query,
    }
    
    return render(request, "adminfile/userlist.html", datas)



def export_users_to_excel(user_list):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

    columns = ['S.N.', 'Full Name', 'Email', 'Phone Number', 'Location']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=row_num, column=col_num, value=column_title)

    for index, user in enumerate(user_list, start=1):
        row_num += 1
        worksheet.cell(row=row_num, column=1, value=index)
        worksheet.cell(row=row_num, column=2, value=user.full_name)
        worksheet.cell(row=row_num, column=3, value=user.user.email)
        worksheet.cell(row=row_num, column=4, value=user.phone)
        worksheet.cell(row=row_num, column=5, value=user.location)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'
    workbook.save(response)
    return response



@login_required
def AddTeam(request):
        if request.method == "POST":
            form = AddteamForm(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                messages.success(request, "Member is saved successfully!")
                return redirect('team_form')
            else:
                messages.error(request, "There was an error saving the product.")
        else:
            form = AddteamForm()

        return render(request, 'adminfile/add-team.html', {'form': form})

@login_required
def delete_team(request, team_id):
    team = get_object_or_404(Add_Team, id=team_id)
    team.delete()
    messages.success(request, 'Member has been deleted successfully.')
    return redirect('admin_dashboard')


@login_required
def delete_selected_teams(request):
    if request.method == 'POST':
        selected_teams = request.POST.getlist('selected_teams')
        if selected_teams:
            Add_Team.objects.filter(id__in=selected_teams).delete()
            messages.success(request, 'Selected members have been deleted successfully.')
        else:
            messages.error(request, 'No members were selected for deletion.')
    return redirect('admin_dashboard')

@login_required
def edit_restaurant_info(request):
    info = RestaurantInfo.objects.first()

    if request.method == 'POST':
        form = RestaurantInfoForm(request.POST, instance=info)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard') 
    else:
        form = RestaurantInfoForm(instance=info)

    return render(request, 'adminfile/about-us-content.html', {'form': form})